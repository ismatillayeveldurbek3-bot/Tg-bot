import os
import re
import csv
import json
import asyncio
import logging
import sqlite3
from functools import wraps
from html import escape
from pathlib import Path
from datetime import datetime, timezone, timedelta
from typing import Optional, Iterable

from aiogram import Bot, Dispatcher, F
from aiogram.enums import ParseMode, ChatMemberStatus
from aiogram.filters import Command, StateFilter
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import (
    Message, CallbackQuery, InlineKeyboardMarkup, InlineKeyboardButton,
    ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove, FSInputFile
)
from aiogram.utils.keyboard import InlineKeyboardBuilder
from aiogram.exceptions import TelegramBadRequest, TelegramForbiddenError
from aiogram.client.default import DefaultBotProperties

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except Exception:  # pragma: no cover
    Workbook = None

try:
    from apscheduler.schedulers.asyncio import AsyncIOScheduler
except Exception:  # pragma: no cover
    AsyncIOScheduler = None

# =====================================================
# CONFIG
# =====================================================
BOT_TOKEN = os.getenv("BOT_TOKEN")
if not BOT_TOKEN:
    raise RuntimeError("BOT_TOKEN env o'zgaruvchisi topilmadi.")

CHANNEL_USERNAME = os.getenv("CHANNEL_USERNAME", "@Qashqadaryo_PMM")
CHANNEL_URL = os.getenv("CHANNEL_URL", "https://t.me/Qashqadaryo_PMM")
FACEBOOK_URL = os.getenv("FACEBOOK_URL", "https://www.facebook.com/share/1E4ZVePTh4/")
INSTAGRAM_URL = os.getenv("INSTAGRAM_URL", "https://www.instagram.com/pedagogikmahorat")
ADMIN_IDS = [int(x) for x in os.getenv("ADMIN_IDS", "5298063089,7361393654").replace(";", ",").split(",") if x.strip().isdigit()]

DATA_DIR = Path(os.getenv("DATA_DIR", "/app/data"))
DATA_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = DATA_DIR / "votes.db"
BACKUP_DIR = DATA_DIR / "backups"
BACKUP_DIR.mkdir(parents=True, exist_ok=True)
UZ_TZ = timezone(timedelta(hours=5), name="Asia/Tashkent")

logging.basicConfig(level=os.getenv("LOG_LEVEL", "INFO"), format="%(asctime)s | %(levelname)s | %(name)s | %(message)s")
log = logging.getLogger("rating-bot")

# =====================================================
# INITIAL DATA / BACKWARD COMPATIBILITY
# =====================================================
SUBJECTS = {
    "s1": {"name": "Tillarni o'qitish metodikasi", "old_key": "tillarni_oqitish_metodikasi", "teachers": {
        "tom_1": "Norov Otajon Shomurodovich", "tom_2": "Abdixolikov Abdulazizxon Abduvohob o'g'li", "tom_3": "Azimova Nigora Anvar qizi", "tom_4": "Abatov Doston Ro'zimurod o'g'li", "tom_5": "Jalilova Komila Abdullayevna", "tom_6": "Oqboyeva Zulfiya Bobonazarovna", "tom_7": "Sevastyanova Nadejda Aleksandrovna", "tom_8": "Xidirova Feruza To'rayevna", "tom_9": "Ergasheva Dilorom Muradilloyevna"}},
    "s2": {"name": "Pedagogika, psixologiya va ta'lim menejmenti", "old_key": "pedagogika_psixologiya_va_talim_menejmenti", "teachers": {
        "pptm_1": "Umarov Lutfillo Murodilloyevich", "pptm_2": "Baratova Nasiba Turobovna", "pptm_3": "Bekmurodova Dilnoza Pirimovna", "pptm_4": "Meyliyev Lobar Nurmatovna", "pptm_5": "Ochilov Og'abek Narzullayevich", "pptm_6": "Shoniyozova Dilafruz Sabirovna", "pptm_7": "Yaratov Xamidjon Muxtorovich", "pptm_8": "Nazarov Asliddin Faxriddin o'g'li", "pptm_9": "Ergasheva Dilafruz Ergamqulovna", "pptm_10": "Soatov Asadulloh Jabborovich"}},
    "s3": {"name": "Aniq va tabiiy fanlar", "old_key": "aniq_va_tabiiy_fanlar", "teachers": {
        "atf_1": "Jobborov Farhod Bo'riyevich", "atf_2": "Karimova Habiba Abduraxmonovna", "atf_3": "Quldoshova Maftuna Jumanzar qizi", "atf_4": "Mallaev Xamro Ro'ziboyevich", "atf_5": "Mamatov Bekzod Farxotovich", "atf_6": "Pardaeva Muqaddas Zafar qizi", "atf_7": "Parmanov Jahongir Rayhonovich", "atf_8": "Rahmatullayev Erkin Shokirovich", "atf_9": "Suyarov Zoir Shojmardonovich", "atf_10": "Tursunova Maftuna Sulton qizi", "atf_11": "Umarov Ibrohimxon Norxuja o'g'li", "atf_12": "Chariev Rashid Ravshanovich", "atf_13": "Elmurodov Sherdil Ergashyevich", "atf_14": "Eshmonov Laziz Norxo'rja o'g'li", "atf_15": "Karaeva Dilfuzaxon Mamasharipovna", "atf_16": "Salomova Madina Sodiq qizi"}},
    "s4": {"name": "Amaliy va ijtimoiy fanlar", "old_key": "amaliy_va_ijtimoiy_fanlar", "teachers": {
        "aif_1": "Yo'ldashev Bekmirza Elmurodovich", "aif_2": "Jabboborov Laziz Hamza o'g'li", "aif_3": "Nurmatov Samandar Fayratovich", "aif_4": "Batoshov Inatillo Kungirovich", "aif_5": "Rajabov Ruslan Bozorovich", "aif_6": "Sanaev Azamat Alponovich", "aif_7": "Shamsiev Jahongir Qulmurod o'g'li", "aif_8": "Xudoyberdiev Axrorboy Nabi o'g'li", "aif_9": "Xasanova Gulnora Qorshanbiyevna", "aif_10": "Eshnazarova Maziya Allanazarovna"}},
    "s5": {"name": "Maktabgacha, boshlang'ich va maxsus ta'lim", "old_key": "maktabgacha_boshlangich_va_maxsus_talim", "teachers": {
        "mbmt_1": "Irisova Sayyora Rajabovna", "mbmt_2": "Azizova Dilnoz Yo'ldoshevna", "mbmt_3": "G'oyimov Umar Eshmurodovich", "mbmt_4": "Ziyotova Madina Mansur qizi", "mbmt_5": "Karimova Umida Sharopovna", "mbmt_6": "Qarshiyeva Guzal Alimardonovna", "mbmt_7": "Qurbanova Xusnora Xudoyberdi qizi", "mbmt_8": "Rajabova Xurshida Hakimovna", "mbmt_9": "Razzaqova Dilnoza Akramovna", "mbmt_10": "Sadinova Marjona Akmal qizi", "mbmt_11": "Shaxmurodova Dilxaxon Almardanovna", "mbmt_12": "Ergasheva Xusniya Mirzoxid qizi", "mbmt_13": "Zaripova Muslima Qurbonovna"}},
}
OLD_TO_NEW_SUBJECT = {v["old_key"]: k for k, v in SUBJECTS.items()}

conn = sqlite3.connect(DB_PATH, check_same_thread=False)
conn.row_factory = sqlite3.Row
db_lock = asyncio.Lock()
bot = Bot(BOT_TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
dp = Dispatcher(storage=MemoryStorage())

# =====================================================
# FSM
# =====================================================
class Register(StatesGroup):
    first_name = State(); last_name = State(); phone = State(); after = State()
class ComplaintFSM(StatesGroup):
    department = State(); teacher = State(); text = State()
class SuggestionFSM(StatesGroup):
    text = State()
class TeacherFSM(StatesGroup):
    add_department = State(); add_name = State(); edit_name = State(); student_count = State(); search = State()
class TeacherEditFSM(StatesGroup):
    name = State(); student_count = State()
class DepartmentFSM(StatesGroup):
    add_name = State(); edit_name = State(); sort_order = State()
class AdminFSM(StatesGroup):
    broadcast = State(); admin_add = State(); admin_remove = State()

# =====================================================
# HELPERS
# =====================================================
def now_str() -> str: return datetime.now(UZ_TZ).strftime("%Y-%m-%d %H:%M:%S")
def is_admin(uid: int) -> bool: return uid in ADMIN_IDS or str(uid) in get_setting("extra_admins", "").split(",")
def safe(s: Optional[str]) -> str: return escape(s or "")
def slug(text: str, prefix: str = "id") -> str:
    # Latin/Cyrillic/Uzbek apostroflar bilan kelgan ismlarni xavfsiz keyga aylantiradi
    base = re.sub(r"[^a-zA-Z0-9_]+", "_", (text or "").lower()).strip("_")[:24]
    return f"{prefix}_{base or int(datetime.now().timestamp())}_{int(datetime.now().timestamp()*1000)%100000}"
def setting_bool(key: str, default="1") -> bool: return get_setting(key, default) == "1"

async def answer_cb(call: CallbackQuery, text: str = ""):
    try: await call.answer(text)
    except TelegramBadRequest: pass

async def edit_or_send(target, text: str, reply_markup=None):
    if isinstance(target, CallbackQuery):
        try:
            await target.message.edit_text(text, reply_markup=reply_markup)
        except TelegramBadRequest:
            await target.message.answer(text, reply_markup=reply_markup)
        except Exception:
            log.exception('edit_or_send failed, sending new message')
            await target.message.answer(text, reply_markup=reply_markup)
    else:
        await target.answer(text, reply_markup=reply_markup)

def cancel_kb() -> ReplyKeyboardMarkup:
    return ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text='❌ Bekor qilish')]], resize_keyboard=True, one_time_keyboard=True)

async def ask_input(target, text: str):
    # Ma'lumot kiritish talab qilingan barcha joylarda oddiy reply keyboard bilan bekor qilish tugmasini ko'rsatadi.
    if isinstance(target, CallbackQuery):
        try:
            await target.message.answer(text, reply_markup=cancel_kb())
        except Exception:
            log.exception('ask_input failed')
    else:
        await target.answer(text, reply_markup=cancel_kb())

async def notify_admins(text: str):
    # Admin xabarnomalari foydalanuvchi flowini qotirib qo'ymasligi uchun alohida xavfsiz yuboriladi.
    for admin_id in ADMIN_IDS:
        try:
            await bot.send_message(admin_id, text[:3900])
        except Exception:
            log.exception('Admin notification failed: %s', admin_id)

# =====================================================
# DATABASE + MIGRATIONS
# =====================================================
def col_exists(table: str, col: str) -> bool:
    cur = conn.execute(f"PRAGMA table_info({table})")
    return any(r[1] == col for r in cur.fetchall())

def table_exists(table: str) -> bool:
    return conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table,)).fetchone() is not None

def get_setting(key: str, default: str = "") -> str:
    row = conn.execute("SELECT value FROM settings WHERE key=?", (key,)).fetchone() if table_exists("settings") else None
    return row[0] if row else default

def set_setting(key: str, value: str):
    conn.execute("INSERT INTO settings(key,value) VALUES(?,?) ON CONFLICT(key) DO UPDATE SET value=excluded.value", (key, value)); conn.commit()

def init_db():
    c = conn.cursor()
    c.execute("CREATE TABLE IF NOT EXISTS settings(key TEXT PRIMARY KEY, value TEXT)")
    c.execute("CREATE TABLE IF NOT EXISTS user_prefs(user_id INTEGER PRIMARY KEY, script TEXT DEFAULT 'latin', access_granted INTEGER DEFAULT 0)")
    c.execute("""CREATE TABLE IF NOT EXISTS users(
        telegram_id INTEGER PRIMARY KEY, first_name TEXT, last_name TEXT, fullname TEXT,
        username TEXT, phone TEXT, registered_at TEXT, updated_at TEXT, is_blocked INTEGER DEFAULT 0)""")
    c.execute("""CREATE TABLE IF NOT EXISTS departments(
        department_key TEXT PRIMARY KEY, name TEXT NOT NULL, sort_order INTEGER DEFAULT 0, is_active INTEGER DEFAULT 1, created_at TEXT)""")
    c.execute("""CREATE TABLE IF NOT EXISTS teachers(
        teacher_key TEXT NOT NULL, department_key TEXT NOT NULL, name TEXT NOT NULL,
        student_count INTEGER DEFAULT 0, is_active INTEGER DEFAULT 1, created_at TEXT,
        PRIMARY KEY(department_key, teacher_key))""")
    c.execute("""CREATE TABLE IF NOT EXISTS ratings(
        id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL, department_key TEXT NOT NULL,
        teacher_key TEXT NOT NULL, rating INTEGER NOT NULL CHECK(rating BETWEEN 1 AND 5),
        fullname TEXT, phone TEXT, username TEXT, rated_at TEXT,
        UNIQUE(user_id, department_key, teacher_key))""")
    c.execute("""CREATE TABLE IF NOT EXISTS complaints(
        id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL, teacher_id TEXT,
        department_key TEXT, teacher_key TEXT, fullname TEXT, phone TEXT, username TEXT,
        complaint_text TEXT, message_text TEXT, created_at TEXT, status TEXT DEFAULT 'Yangi')""")
    c.execute("""CREATE TABLE IF NOT EXISTS suggestions(
        id INTEGER PRIMARY KEY AUTOINCREMENT, user_id INTEGER NOT NULL, fullname TEXT, phone TEXT, username TEXT,
        suggestion_text TEXT NOT NULL, created_at TEXT, status TEXT DEFAULT 'Yangi')""")
    # legacy tables preserved
    c.execute("""CREATE TABLE IF NOT EXISTS votes(
        user_id INTEGER PRIMARY KEY, full_name TEXT, username TEXT, subject_key TEXT NOT NULL, teacher_key TEXT NOT NULL, voted_at TEXT)""")
    conn.commit()

    for col, ddl in [("student_count", "ALTER TABLE teachers ADD COLUMN student_count INTEGER DEFAULT 0"), ("is_active", "ALTER TABLE teachers ADD COLUMN is_active INTEGER DEFAULT 1")]:
        if not col_exists("teachers", col): c.execute(ddl)
    for col, ddl in [("phone", "ALTER TABLE users ADD COLUMN phone TEXT"), ("is_blocked", "ALTER TABLE users ADD COLUMN is_blocked INTEGER DEFAULT 0")]:
        if not col_exists("users", col): c.execute(ddl)
    for col, ddl in [("user_id", "ALTER TABLE complaints ADD COLUMN user_id INTEGER"), ("teacher_id", "ALTER TABLE complaints ADD COLUMN teacher_id TEXT"), ("department_key", "ALTER TABLE complaints ADD COLUMN department_key TEXT"), ("teacher_key", "ALTER TABLE complaints ADD COLUMN teacher_key TEXT"), ("fullname", "ALTER TABLE complaints ADD COLUMN fullname TEXT"), ("phone", "ALTER TABLE complaints ADD COLUMN phone TEXT"), ("username", "ALTER TABLE complaints ADD COLUMN username TEXT"), ("complaint_text", "ALTER TABLE complaints ADD COLUMN complaint_text TEXT"), ("message_text", "ALTER TABLE complaints ADD COLUMN message_text TEXT"), ("created_at", "ALTER TABLE complaints ADD COLUMN created_at TEXT"), ("status", "ALTER TABLE complaints ADD COLUMN status TEXT DEFAULT 'Yangi'")]:
        if not col_exists("complaints", col): c.execute(ddl)
    for col, ddl in [("fullname", "ALTER TABLE suggestions ADD COLUMN fullname TEXT"), ("phone", "ALTER TABLE suggestions ADD COLUMN phone TEXT"), ("username", "ALTER TABLE suggestions ADD COLUMN username TEXT"), ("suggestion_text", "ALTER TABLE suggestions ADD COLUMN suggestion_text TEXT"), ("created_at", "ALTER TABLE suggestions ADD COLUMN created_at TEXT"), ("status", "ALTER TABLE suggestions ADD COLUMN status TEXT DEFAULT 'Yangi'")]:
        if not col_exists("suggestions", col): c.execute(ddl)
    conn.commit()

    # seed departments/teachers without overwriting admin changes
    if conn.execute("SELECT COUNT(*) FROM departments").fetchone()[0] == 0:
        for i, (dkey, data) in enumerate(SUBJECTS.items()):
            c.execute("INSERT OR IGNORE INTO departments(department_key,name,sort_order,created_at) VALUES(?,?,?,?)", (dkey, data["name"], i, now_str()))
            for tkey, tname in data["teachers"].items():
                c.execute("INSERT OR IGNORE INTO teachers(teacher_key,department_key,name,created_at) VALUES(?,?,?,?)", (tkey, dkey, tname, now_str()))
    # migrate old db_subjects/db_teachers if they exist
    if table_exists("db_subjects"):
        for r in conn.execute("SELECT subject_key, subject_name, sort_order FROM db_subjects"):
            dkey = OLD_TO_NEW_SUBJECT.get(r[0], r[0])
            c.execute("INSERT OR IGNORE INTO departments(department_key,name,sort_order,created_at) VALUES(?,?,?,?)", (dkey, r[1], r[2] or 0, now_str()))
    if table_exists("db_teachers"):
        for r in conn.execute("SELECT teacher_key, subject_key, teacher_name FROM db_teachers"):
            dkey = OLD_TO_NEW_SUBJECT.get(r[1], r[1])
            c.execute("INSERT OR IGNORE INTO teachers(teacher_key,department_key,name,created_at) VALUES(?,?,?,?)", (r[0], dkey, r[2], now_str()))
    # migrate old user names from votes/teacher_ratings
    for table in ["votes", "teacher_ratings"]:
        if table_exists(table):
            cols = [x[1] for x in conn.execute(f"PRAGMA table_info({table})")]
            if "user_id" in cols:
                name_col = "full_name" if "full_name" in cols else "fullname" if "fullname" in cols else None
                uname_col = "username" if "username" in cols else None
                select = f"SELECT user_id{','+name_col if name_col else ''}{','+uname_col if uname_col else ''} FROM {table}"
                for row in conn.execute(select):
                    uid = row[0]; full = row[1] if name_col else ""; uname = row[2] if name_col and uname_col else (row[1] if uname_col and not name_col else "")
                    c.execute("INSERT OR IGNORE INTO users(telegram_id, fullname, username, registered_at, updated_at) VALUES(?,?,?,?,?)", (uid, full, uname, now_str(), now_str()))
    # migrate like/dislike table into 1-5 ratings
    if table_exists("teacher_ratings"):
        cols = [x[1] for x in conn.execute("PRAGMA table_info(teacher_ratings)")]
        if "rating" in cols:
            for r in conn.execute("SELECT * FROM teacher_ratings"):
                dkey = OLD_TO_NEW_SUBJECT.get(r["subject_key"], r["subject_key"]) if "subject_key" in r.keys() else r["department_key"]
                val = r["rating"]
                try: iv = int(val)
                except Exception: iv = 5 if val == "like" else 1
                c.execute("""INSERT OR IGNORE INTO ratings(user_id,department_key,teacher_key,rating,fullname,username,rated_at)
                             VALUES(?,?,?,?,?,?,?)""", (r["user_id"], dkey, r["teacher_key"], iv, r["full_name"] if "full_name" in r.keys() else "", r["username"] if "username" in r.keys() else "", r["rated_at"] if "rated_at" in r.keys() else now_str()))
    # normalize old subject keys
    for old, new in OLD_TO_NEW_SUBJECT.items():
        c.execute("UPDATE votes SET subject_key=? WHERE subject_key=?", (new, old))
        c.execute("UPDATE ratings SET department_key=? WHERE department_key=?", (new, old))
    # indexes
    indexes = [
        "CREATE INDEX IF NOT EXISTS idx_ratings_teacher ON ratings(department_key,teacher_key)",
        "CREATE INDEX IF NOT EXISTS idx_ratings_user ON ratings(user_id)",
        "CREATE INDEX IF NOT EXISTS idx_complaints_teacher ON complaints(department_key,teacher_key)",
        "CREATE INDEX IF NOT EXISTS idx_complaints_status ON complaints(status)",
        "CREATE INDEX IF NOT EXISTS idx_suggestions_status ON suggestions(status)",
        "CREATE INDEX IF NOT EXISTS idx_suggestions_user ON suggestions(user_id)",
        "CREATE INDEX IF NOT EXISTS idx_teachers_department ON teachers(department_key,is_active)",
    ]
    for q in indexes: c.execute(q)
    if get_setting("voting_open", "") == "": set_setting("voting_open", "1")
    if get_setting("daily_backup", "") == "": set_setting("daily_backup", "1")
    conn.commit()

# =====================================================
# DATA ACCESS
# =====================================================
def ensure_user_obj(m: Message):
    u = m.from_user
    full = " ".join([x for x in [u.first_name, u.last_name] if x])
    conn.execute("""INSERT INTO users(telegram_id,first_name,last_name,fullname,username,registered_at,updated_at)
        VALUES(?,?,?,?,?,?,?) ON CONFLICT(telegram_id) DO UPDATE SET username=excluded.username, updated_at=excluded.updated_at""",
        (u.id, u.first_name, u.last_name, full, u.username, now_str(), now_str()))
    conn.execute("INSERT OR IGNORE INTO user_prefs(user_id) VALUES(?)", (u.id,)); conn.commit()

def user_registered(uid: int) -> bool:
    r = conn.execute("SELECT first_name,last_name,phone FROM users WHERE telegram_id=?", (uid,)).fetchone()
    return bool(r and r[0] and r[1] and r[2])

def get_user(uid: int): return conn.execute("SELECT * FROM users WHERE telegram_id=?", (uid,)).fetchone()
def departments(active=True):
    q = "SELECT * FROM departments" + (" WHERE is_active=1" if active else "") + " ORDER BY sort_order,name"
    return conn.execute(q).fetchall()
def teachers(dkey: str, active=True):
    q = "SELECT * FROM teachers WHERE department_key=?" + (" AND is_active=1" if active else "") + " ORDER BY name"
    return conn.execute(q, (dkey,)).fetchall()
def all_teachers(active=True):
    q = """SELECT t.*, d.name AS department_name FROM teachers t JOIN departments d ON d.department_key=t.department_key"""
    if active: q += " WHERE t.is_active=1 AND d.is_active=1"
    q += " ORDER BY d.sort_order,t.name"
    return conn.execute(q).fetchall()
def dep_name(dkey):
    r = conn.execute("SELECT name FROM departments WHERE department_key=?", (dkey,)).fetchone(); return r[0] if r else dkey
def teacher_name(dkey,tkey):
    r = conn.execute("SELECT name FROM teachers WHERE department_key=? AND teacher_key=?", (dkey,tkey)).fetchone(); return r[0] if r else tkey

def save_rating(uid:int, dkey:str, tkey:str, rating:int) -> bool:
    # Talab bo‘yicha: bir foydalanuvchi bir o‘qituvchini faqat bir marta baholaydi.
    exists = conn.execute("SELECT id FROM ratings WHERE user_id=? AND department_key=? AND teacher_key=?", (uid, dkey, tkey)).fetchone()
    if exists:
        return False
    u=get_user(uid); full=(u["fullname"] if u else "") or " ".join(filter(None,[u["first_name"] if u else "", u["last_name"] if u else ""]))
    conn.execute("""INSERT INTO ratings(user_id,department_key,teacher_key,rating,fullname,phone,username,rated_at)
        VALUES(?,?,?,?,?,?,?,?)""",
        (uid,dkey,tkey,rating,full,u["phone"] if u else "",u["username"] if u else "",now_str()))
    conn.commit()
    return True

def teacher_stats(dkey:str,tkey:str):
    t=conn.execute("SELECT * FROM teachers WHERE department_key=? AND teacher_key=?",(dkey,tkey)).fetchone()
    rows=conn.execute("SELECT rating,COUNT(*) c FROM ratings WHERE department_key=? AND teacher_key=? GROUP BY rating",(dkey,tkey)).fetchall()
    dist={i:0 for i in range(1,6)}
    total=0; weighted=0
    for r in rows:
        dist[int(r[0])]=r[1]; total+=r[1]; weighted+=int(r[0])*r[1]
    avg=weighted/total if total else 0
    sc=(t["student_count"] or 0) if t else 0
    participation=(total/sc*100) if sc else 0
    final=avg*(total/sc) if sc else 0
    return {"avg":avg,"total":total,"student_count":sc,"participation":participation,"final":final,"dist":dist}

def department_stats(dkey:str):
    ts=teachers(dkey)
    total_teachers=len(ts)
    total_votes=0; weighted_sum=0; student_sum=0; finals=[]
    for t in ts:
        st=teacher_stats(dkey,t["teacher_key"])
        total_votes += st["total"]
        weighted_sum += st["avg"]*st["total"]
        student_sum += st["student_count"]
        finals.append(st["final"])
    avg=weighted_sum/total_votes if total_votes else 0
    participation=total_votes/student_sum*100 if student_sum else 0
    final=sum(finals)/len(finals) if finals else 0
    return {"teachers":total_teachers,"votes":total_votes,"avg":avg,"participation":participation,"final":final,"student_sum":student_sum}

# =====================================================
# KEYBOARDS
# =====================================================
def ik(rows: list[list[tuple[str,str]]]) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text=t, callback_data=d) for t,d in row] for row in rows])
def main_kb(): return ik([[('📝 Baholash','rate:start')],[('📢 Shikoyat','complaint:start'),('💡 Taklif','suggestion:start')],[('🏆 Reyting','rating:top'),('ℹ️ Ma’lumot','info')]])
def admin_kb(): return ik([[('📊 Dashboard','admin:dash')],[('👨‍🏫 O‘qituvchilar','admin:teachers'),('🏆 Reyting','admin:rating')],[('📢 Shikoyatlar','admin:complaints'),('💡 Takliflar','admin:suggestions')],[('📁 Export / Backup','admin:export'),('⚙️ Sozlamalar','admin:settings')]])
def deps_kb(prefix:str, back='home'):
    rows=[[ (d['name'][:40], f'{prefix}:{d["department_key"]}') ] for d in departments()]
    rows.append([('⬅️ Orqaga',back)]); return ik(rows)
def teachers_kb(dkey:str,prefix:str,back:str):
    rows=[[ (t['name'][:45], f'{prefix}:{dkey}:{t["teacher_key"]}') ] for t in teachers(dkey)]
    rows.append([('⬅️ Orqaga',back)]); return ik(rows)
def rating_stars_kb(dkey,tkey):
    return ik([[('⭐ 1',f'rate:save:{dkey}:{tkey}:1'),('⭐⭐ 2',f'rate:save:{dkey}:{tkey}:2')],[('⭐⭐⭐ 3',f'rate:save:{dkey}:{tkey}:3')],[('⭐⭐⭐⭐ 4',f'rate:save:{dkey}:{tkey}:4'),('⭐⭐⭐⭐⭐ 5',f'rate:save:{dkey}:{tkey}:5')],[('⬅️ O‘qituvchilar',f'rate:dep:{dkey}')]])
def phone_kb(): return ReplyKeyboardMarkup(keyboard=[[KeyboardButton(text='📱 Telefon raqamni yuborish', request_contact=True)], [KeyboardButton(text='❌ Bekor qilish')]], resize_keyboard=True, one_time_keyboard=True)

def admin_teachers_kb(): return ik([[('➕ Qo‘shish','tm:add'),('📋 Ro‘yxat','tm:list')],[('🔍 Qidirish','tm:search'),('👥 O‘quvchilar soni','tm:students')],[('🏛 Kafedralar','dep:manage')],[('⬅️ Admin','admin:menu')]])
def admin_settings_kb(): return ik([[('🏛 Kafedralar','set:deps'),('📢 Shikoyatlar','set:complaints')],[('🗳 Baholash','set:voting'),('📊 Reyting','set:rating')],[('🔐 Adminlar','set:admins'),('📦 Backup','set:backup')],[('⬅️ Admin','admin:menu')]])

# =====================================================
# TEXTS
# =====================================================
MOTIVATION = """<b>Sizning fikringiz ta’lim sifatini oshirish uchun muhim.</b>\n\nHar bir berilgan baho o‘qituvchilar faoliyatini xolis tahlil qilish, kuchli jihatlarni aniqlash va rivojlantirilishi kerak bo‘lgan yo‘nalishlarni ko‘rish uchun xizmat qiladi.\n\nIltimos, baholashda adolatli, xolis va mas’uliyatli bo‘ling. Sizning javobingiz umumiy reyting va sifat ko‘rsatkichlariga ta’sir qiladi."""
def home_text(): return "<b>Bosh menyu</b>\nKerakli bo‘limni tanlang."
def info_text(): return "<b>Ma’lumot</b>\n\nBot o‘qituvchilar faoliyatini 1–5 yulduzli tizimda baholash, reytinglarni ko‘rish va shikoyat va taklif yuborish uchun mo‘ljallangan.\n\nFormula: <code>final_score = average_rating × (total_votes / student_count)</code>"

def teacher_stats_text(dkey,tkey):
    st=teacher_stats(dkey,tkey); dist='\n'.join([f"{i}⭐: {st['dist'][i]} ta" for i in range(1,6)])
    return f"<b>O‘qituvchi statistikasi</b>\n\n<b>O‘qituvchi:</b> {safe(teacher_name(dkey,tkey))}\n<b>Kafedra:</b> {safe(dep_name(dkey))}\n\nO‘rtacha baho: <b>{st['avg']:.2f}</b>\nOvozlar soni: <b>{st['total']}</b>\nO‘quvchilar soni: <b>{st['student_count']}</b>\nQatnashuv: <b>{st['participation']:.1f}%</b>\nFinal score: <b>{st['final']:.3f}</b>\n\n<b>Rating distribution</b>\n{dist}"

def top_rating_text(limit=20):
    data=[]
    for t in all_teachers():
        st=teacher_stats(t['department_key'],t['teacher_key'])
        data.append((st['final'],st['avg'],st['total'],t))
    data.sort(key=lambda x:(x[0],x[2],x[1]), reverse=True)
    lines=["🏆 <b>O‘qituvchilar reytingi</b>\n"]
    for i,(final,avg,total,t) in enumerate(data[:limit],1):
        lines.append(f"{i}. <b>{safe(t['name'])}</b>\n{safe(t['department_name'])}\n⭐ {avg:.2f} • {total} baho • score: <b>{final:.3f}</b>")
    return '\n\n'.join(lines)

def departments_rating_text():
    data=[]
    for d in departments():
        st=department_stats(d['department_key']); data.append((st['final'],d,st))
    data.sort(key=lambda x:x[0], reverse=True)
    lines=["🏛 <b>Kafedralar reytingi</b>\n"]
    for i,(score,d,st) in enumerate(data,1):
        lines.append(f"{i}. <b>{safe(d['name'])}</b>\nO‘qituvchi: {st['teachers']} • Baholar: {st['votes']}\nO‘rtacha: {st['avg']:.2f} • Qatnashuv: {st['participation']:.1f}%\nFinal score: <b>{st['final']:.3f}</b>")
    return '\n\n'.join(lines)



def teacher_ranking_rows(limit: int | None = None):
    data=[]
    for t in all_teachers():
        st=teacher_stats(t['department_key'],t['teacher_key'])
        data.append((st['final'], st['avg'], st['total'], t, st))
    data.sort(key=lambda x:(x[0], x[2], x[1]), reverse=True)
    return data[:limit] if limit else data

def department_ranking_rows():
    data=[]
    for d in departments():
        st=department_stats(d['department_key'])
        data.append((st['final'], d, st))
    data.sort(key=lambda x:x[0], reverse=True)
    return data

def admin_teacher_top10_text():
    rows=teacher_ranking_rows(10)
    lines=["🏆 <b>TOP 10 o‘qituvchilar reytingi</b>\n"]
    if not rows:
        return "🏆 <b>TOP 10 o‘qituvchilar reytingi</b>\n\nHozircha ma’lumot yo‘q."
    for i,(final,avg,total,t,st) in enumerate(rows,1):
        lines.append(f"{i}. <b>{safe(t['name'])}</b>\n{safe(t['department_name'])}\n⭐ {avg:.2f} • {total} baho • Qatnashuv: {st['participation']:.1f}% • score: <b>{final:.3f}</b>")
    return '\n\n'.join(lines)

def admin_department_rating_text():
    rows=department_ranking_rows()
    lines=["🏛 <b>Kafedralar reytingi</b>\n<em>Hisob: kafedradagi o‘qituvchilar final score o‘rta arifmetigi.</em>\n"]
    if not rows:
        return "🏛 <b>Kafedralar reytingi</b>\n\nHozircha ma’lumot yo‘q."
    for i,(score,d,st) in enumerate(rows,1):
        lines.append(f"{i}. <b>{safe(d['name'])}</b>\nO‘qituvchilar: {st['teachers']} • Jami baholar: {st['votes']}\nO‘rtacha baho: {st['avg']:.2f} • Qatnashuv: {st['participation']:.1f}%\nFinal: <b>{st['final']:.3f}</b>")
    return '\n\n'.join(lines)

# =====================================================
# EXPORT / BACKUP
# =====================================================
def style_ws(ws):
    if Workbook is None: return
    header_fill=PatternFill("solid", fgColor="1F4E78"); header_font=Font(color="FFFFFF", bold=True)
    thin=Side(style="thin", color="D9E2F3")
    for cell in ws[1]: cell.fill=header_fill; cell.font=header_font; cell.alignment=Alignment(horizontal="center"); cell.border=Border(bottom=thin)
    for col in ws.columns:
        max_len=max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width=min(max(max_len+2,12),45)
    ws.freeze_panes="A2"

def export_all_excel() -> Path:
    if Workbook is None:
        path=DATA_DIR/"export.csv"; return path
    wb=Workbook(); wb.remove(wb.active)
    sheets = {
        "Users": ("SELECT telegram_id,first_name,last_name,fullname,username,phone,registered_at FROM users", ["Telegram ID","Ism","Familiya","FISH","Username","Telefon","Ro‘yxat"]),
        "Departments": ("SELECT department_key,name,sort_order,is_active FROM departments", ["Key","Nomi","Tartib","Active"]),
        "Teachers": ("SELECT department_key,teacher_key,name,student_count,is_active FROM teachers", ["Kafedra","Teacher key","O‘qituvchi","O‘quvchilar soni","Active"]),
        "Ratings": ("SELECT user_id,department_key,teacher_key,rating,fullname,phone,username,rated_at FROM ratings ORDER BY rated_at DESC", ["User ID","Kafedra","Teacher key","Baho","FISH","Telefon","Username","Sana"]),
        "Complaints": ("SELECT id,user_id,department_key,teacher_key,fullname,phone,complaint_text,status,created_at FROM complaints ORDER BY id DESC", ["ID","User ID","Kafedra","Teacher key","FISH","Telefon","Shikoyat","Status","Sana"]),
        "Suggestions": ("SELECT id,user_id,fullname,phone,username,suggestion_text,status,created_at FROM suggestions ORDER BY id DESC", ["ID","User ID","FISH","Telefon","Username","Taklif","Status","Sana"]),
    }
    for name,(query,headers) in sheets.items():
        ws=wb.create_sheet(name); ws.append(headers)
        for r in conn.execute(query): ws.append(list(r))
        style_ws(ws)
    ws=wb.create_sheet("Teacher Stats"); ws.append(["Kafedra","O‘qituvchi","Average","Votes","Students","Participation %","Final score","1⭐","2⭐","3⭐","4⭐","5⭐"])
    for t in all_teachers(active=False):
        st=teacher_stats(t['department_key'],t['teacher_key'])
        ws.append([dep_name(t['department_key']),t['name'],round(st['avg'],2),st['total'],st['student_count'],round(st['participation'],2),round(st['final'],4),st['dist'][1],st['dist'][2],st['dist'][3],st['dist'][4],st['dist'][5]])
    style_ws(ws)
    ws=wb.create_sheet("Teacher Ranking"); ws.append(["O‘rin","Kafedra","O‘qituvchi","Average","Votes","Students","Participation %","Final score"])
    for i,(final,avg,total,t,st) in enumerate(teacher_ranking_rows(),1):
        ws.append([i, dep_name(t['department_key']), t['name'], round(avg,2), total, st['student_count'], round(st['participation'],2), round(final,4)])
    style_ws(ws)
    ws=wb.create_sheet("Department Ranking"); ws.append(["O‘rin","Kafedra","Teachers","Votes","Average","Participation %","Final department score"])
    for i,(score,d,st) in enumerate(department_ranking_rows(),1):
        ws.append([i, d['name'], st['teachers'], st['votes'], round(st['avg'],2), round(st['participation'],2), round(score,4)])
    style_ws(ws)
    path=DATA_DIR/f"full_export_{datetime.now(UZ_TZ).strftime('%Y%m%d_%H%M%S')}.xlsx"; wb.save(path); return path

def export_teacher_ranking_excel() -> Path:
    if Workbook is None:
        path=DATA_DIR/"teacher_ranking.csv"; return path
    wb=Workbook(); ws=wb.active; ws.title="Top Teachers"
    ws.append(["O‘rin","Kafedra","O‘qituvchi","Average","Votes","Students","Participation %","Final score"])
    for i,(final,avg,total,t,st) in enumerate(teacher_ranking_rows(10),1):
        ws.append([i, dep_name(t['department_key']), t['name'], round(avg,2), total, st['student_count'], round(st['participation'],2), round(final,4)])
    style_ws(ws)
    path=DATA_DIR/f"teacher_top10_{datetime.now(UZ_TZ).strftime('%Y%m%d_%H%M%S')}.xlsx"; wb.save(path); return path

def export_department_ranking_excel() -> Path:
    if Workbook is None:
        path=DATA_DIR/"department_ranking.csv"; return path
    wb=Workbook(); ws=wb.active; ws.title="Department Ranking"
    ws.append(["O‘rin","Kafedra","Teachers","Votes","Average","Participation %","Final department score"])
    for i,(score,d,st) in enumerate(department_ranking_rows(),1):
        ws.append([i, d['name'], st['teachers'], st['votes'], round(st['avg'],2), round(st['participation'],2), round(score,4)])
    style_ws(ws)
    path=DATA_DIR/f"department_ranking_{datetime.now(UZ_TZ).strftime('%Y%m%d_%H%M%S')}.xlsx"; wb.save(path); return path


def make_backup() -> Path:
    import zipfile
    export_path=export_all_excel()
    backup=BACKUP_DIR/f"backup_{datetime.now(UZ_TZ).strftime('%Y%m%d_%H%M%S')}.zip"
    conn.commit()
    with zipfile.ZipFile(backup,"w",zipfile.ZIP_DEFLATED) as z:
        if DB_PATH.exists(): z.write(DB_PATH, "votes.db")
        if export_path.exists(): z.write(export_path, export_path.name)
    return backup

# =====================================================
# ACCESS / SUBSCRIPTION
# =====================================================
async def check_sub(uid:int) -> bool:
    if not setting_bool("subscription_required", "1"): return True
    try:
        m=await bot.get_chat_member(CHANNEL_USERNAME, uid)
        return m.status in {ChatMemberStatus.MEMBER, ChatMemberStatus.ADMINISTRATOR, ChatMemberStatus.CREATOR}
    except Exception:
        return False

def sub_kb(): return ik([[('📢 Telegram kanal', 'url:noop')],[('✅ Tekshirish','check_sub')]])

# =====================================================
# USER HANDLERS
# =====================================================
@dp.message(Command("start"))
async def start(m: Message, state:FSMContext):
    await state.clear(); ensure_user_obj(m)
    if not await check_sub(m.from_user.id):
        kb=InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text='📢 Telegram kanal', url=CHANNEL_URL)],[InlineKeyboardButton(text='✅ Tekshirish', callback_data='check_sub')]])
        return await m.answer("Botdan foydalanish uchun Telegram kanalga obuna bo‘ling.", reply_markup=kb)
    await m.answer(home_text(), reply_markup=main_kb())

@dp.callback_query(F.data == 'check_sub')
async def cb_check(call: CallbackQuery):
    await answer_cb(call)
    if await check_sub(call.from_user.id): await edit_or_send(call, home_text(), main_kb())
    else: await call.answer("Avval kanalga obuna bo‘ling.", show_alert=True)

@dp.message(Command("admin"))
async def admin_cmd(m:Message):
    ensure_user_obj(m)
    if not is_admin(m.from_user.id): return await m.answer("Ruxsat yo‘q.")
    await m.answer("<b>Admin panel</b>", reply_markup=admin_kb())

@dp.callback_query(F.data == 'home')
async def home_cb(call:CallbackQuery): await answer_cb(call); await edit_or_send(call, home_text(), main_kb())
@dp.callback_query(F.data == 'info')
async def info_cb(call:CallbackQuery): await answer_cb(call); await edit_or_send(call, info_text(), ik([[('⬅️ Orqaga','home')]]))

@dp.message(StateFilter('*'), F.text == '❌ Bekor qilish')
async def cancel_button(m: Message, state: FSMContext):
    current = await state.get_state()
    data = await state.get_data()
    await m.answer('⬅️ Bir qadam orqaga qaytildi.', reply_markup=ReplyKeyboardRemove())
    if current == Register.phone.state:
        await state.set_state(Register.last_name)
        return await ask_input(m, 'Familiyangizni qaytadan kiriting:')
    if current == Register.last_name.state:
        await state.set_state(Register.first_name)
        return await ask_input(m, 'Ismingizni qaytadan kiriting:')
    if current == Register.first_name.state:
        await state.clear()
        return await m.answer(home_text(), reply_markup=main_kb())
    if current == ComplaintFSM.text.state:
        dkey=data.get('dkey')
        await state.set_state(ComplaintFSM.teacher)
        return await m.answer('O‘qituvchini tanlang:', reply_markup=teachers_kb(dkey,'complaint:teacher','complaint:start') if dkey else deps_kb('complaint:dep','home'))
    if current == SuggestionFSM.text.state:
        await state.clear()
        return await m.answer(home_text(), reply_markup=main_kb())
    if current == TeacherFSM.add_name.state:
        await state.set_state(TeacherFSM.add_department)
        return await m.answer('Qaysi kafedraga qo‘shiladi?', reply_markup=deps_kb('tm:adddep','admin:teachers'))
    if current == TeacherFSM.student_count.state:
        dkey=data.get('dkey'); tkey=data.get('tkey')
        await state.clear()
        if dkey and tkey:
            return await m.answer(teacher_stats_text(dkey,tkey), reply_markup=ik([[('⬅️ O‘qituvchilar',f'tm:listdep:{dkey}')],[('⬅️ Admin','admin:menu')]]))
        return await m.answer('👨‍🏫 <b>O‘qituvchilar</b>', reply_markup=admin_teachers_kb())
    if current == TeacherFSM.search.state:
        await state.clear()
        return await m.answer('👨‍🏫 <b>O‘qituvchilar</b>', reply_markup=admin_teachers_kb())
    if current in (TeacherEditFSM.name.state, TeacherEditFSM.student_count.state):
        dkey=data.get('dkey'); tkey=data.get('tkey')
        await state.clear()
        if dkey and tkey:
            return await m.answer(teacher_stats_text(dkey,tkey), reply_markup=ik([[('✏️ Ism',f'tm:editname:{dkey}:{tkey}'),('👥 O‘quvchi',f'tm:editstudents:{dkey}:{tkey}')],[('⬅️ Ro‘yxat',f'tm:listdep:{dkey}')]]))
        return await m.answer('👨‍🏫 <b>O‘qituvchilar</b>', reply_markup=admin_teachers_kb())
    if current in (DepartmentFSM.add_name.state, DepartmentFSM.edit_name.state, DepartmentFSM.sort_order.state):
        await state.clear()
        return await m.answer('🏛 <b>Kafedralar boshqaruvi</b>', reply_markup=ik([[('➕ Qo‘shish','dep:add'),('📋 Ro‘yxat','dep:list')],[('⬅️ Admin','admin:menu')]]))
    if current in (AdminFSM.admin_add.state, AdminFSM.admin_remove.state):
        await state.clear()
        return await m.answer('🔐 Adminlar sozlamasi', reply_markup=ik([[('➕ Admin qo‘shish','admin:add'),('➖ Admin olish','admin:remove')],[('⬅️ Sozlamalar','admin:settings')]]))
    await state.clear()
    if is_admin(m.from_user.id):
        await m.answer('<b>Admin panel</b>', reply_markup=admin_kb())
    else:
        await m.answer(home_text(), reply_markup=main_kb())

@dp.message(Register.first_name)
async def reg_first(m:Message,state:FSMContext):
    first=(m.text or "").strip()
    if len(first) < 2: return await m.answer("Ism juda qisqa. Qaytadan kiriting:", reply_markup=cancel_kb())
    await state.update_data(first_name=first); await state.set_state(Register.last_name); await ask_input(m, "Familiyangizni kiriting:")
@dp.message(Register.last_name)
async def reg_last(m:Message,state:FSMContext):
    last=(m.text or "").strip()
    if len(last) < 2: return await m.answer("Familiya juda qisqa. Qaytadan kiriting:", reply_markup=cancel_kb())
    await state.update_data(last_name=last); await state.set_state(Register.phone); await m.answer("Telefon raqamingizni contact button orqali yuboring:", reply_markup=phone_kb())
@dp.message(Register.phone)
async def reg_phone(m:Message,state:FSMContext):
    if not m.contact or m.contact.user_id != m.from_user.id: return await m.answer("Iltimos, pastdagi contact tugmasi orqali o‘z telefon raqamingizni yuboring.", reply_markup=phone_kb())
    data=await state.get_data(); full=f"{data['first_name']} {data['last_name']}"
    conn.execute("""INSERT INTO users(telegram_id,first_name,last_name,fullname,username,phone,registered_at,updated_at)
        VALUES(?,?,?,?,?,?,?,?) ON CONFLICT(telegram_id) DO UPDATE SET first_name=excluded.first_name,last_name=excluded.last_name,fullname=excluded.fullname,username=excluded.username,phone=excluded.phone,updated_at=excluded.updated_at""",
        (m.from_user.id,data['first_name'],data['last_name'],full,m.from_user.username,m.contact.phone_number,now_str(),now_str()))
    conn.commit(); after=data.get('after','rate') ; await state.clear()
    await m.answer("Ro‘yxatdan o‘tish yakunlandi.", reply_markup=ReplyKeyboardRemove())
    if after=='complaint': await m.answer("Kafedrani tanlang:", reply_markup=deps_kb('complaint:dep','home'))
    elif after=='suggestion':
        await state.set_state(SuggestionFSM.text)
        await ask_input(m, "💡 Taklifingizni yozing. Matn aniq va lo‘nda bo‘lsa, ko‘rib chiqish oson bo‘ladi:")
    else: await m.answer(MOTIVATION, reply_markup=deps_kb('rate:dep','home'))

@dp.callback_query(F.data == 'rate:start')
async def rate_start(call:CallbackQuery,state:FSMContext):
    await answer_cb(call)
    await state.clear()
    if not await check_sub(call.from_user.id):
        return await call.answer('Avval kanalga obuna bo‘ling.', show_alert=True)
    if not setting_bool('voting_open','1'): return await call.answer('Baholash hozircha yopilgan.', show_alert=True)
    if not user_registered(call.from_user.id): return await start_registration(call,state,'rate')
    await edit_or_send(call, MOTIVATION, deps_kb('rate:dep','home'))
@dp.callback_query(F.data.startswith('rate:dep:'))
async def rate_dep(call:CallbackQuery):
    await answer_cb(call); dkey=call.data.split(':',2)[2]
    await edit_or_send(call, f"<b>{safe(dep_name(dkey))}</b>\nO‘qituvchini tanlang:", teachers_kb(dkey,'rate:teacher','rate:start'))
@dp.callback_query(F.data.startswith('rate:teacher:'))
async def rate_teacher(call:CallbackQuery):
    await answer_cb(call); _,_,dkey,tkey=call.data.split(':',3)
    old=conn.execute("SELECT rating FROM ratings WHERE user_id=? AND department_key=? AND teacher_key=?",(call.from_user.id,dkey,tkey)).fetchone()
    if old:
        return await edit_or_send(call, f"<b>{safe(teacher_name(dkey,tkey))}</b>\n\nSiz bu o‘qituvchini avval baholagansiz: <b>{old[0]}⭐</b>", ik([[('⬅️ O‘qituvchilar',f'rate:dep:{dkey}')],[('🏠 Menyu','home')]]))
    txt=f"<b>{safe(teacher_name(dkey,tkey))}</b>\n\nBahoni tanlang."
    await edit_or_send(call, txt, rating_stars_kb(dkey,tkey))
@dp.callback_query(F.data.startswith('rate:save:'))
async def rate_save(call:CallbackQuery):
    await answer_cb(call); _,_,dkey,tkey,val=call.data.split(':',4)
    if not user_registered(call.from_user.id): return await call.answer('Avval ro‘yxatdan o‘ting.', show_alert=True)
    ok = save_rating(call.from_user.id,dkey,tkey,int(val))
    if not ok:
        return await call.answer('Siz bu o‘qituvchini avval baholagansiz.', show_alert=True)
    await edit_or_send(call, f"✅ Baho saqlandi: <b>{val}⭐</b>\n\n{safe(teacher_name(dkey,tkey))}", ik([[('Yana baholash','rate:start'),('🏆 Reyting','rating:top')],[('🏠 Menyu','home')]]))

@dp.callback_query(F.data == 'rating:top')
async def rating_top(call:CallbackQuery): await answer_cb(call); await edit_or_send(call, top_rating_text(), ik([[('🏛 Kafedralar','dept_rating:public')],[('⬅️ Orqaga','home')]]))
@dp.callback_query(F.data == 'dept_rating:public')
async def public_deps(call:CallbackQuery): await answer_cb(call); await edit_or_send(call, departments_rating_text(), ik([[('⬅️ Reyting','rating:top')]]))

@dp.callback_query(F.data == 'complaint:start')
async def complaint_start(call:CallbackQuery,state:FSMContext):
    await answer_cb(call)
    await state.clear()
    if not user_registered(call.from_user.id): return await start_registration(call,state,'complaint')
    await state.set_state(ComplaintFSM.department); await edit_or_send(call,"Kafedrani tanlang:", deps_kb('complaint:dep','home'))
@dp.callback_query(F.data.startswith('complaint:dep:'))
async def complaint_dep(call:CallbackQuery,state:FSMContext):
    await answer_cb(call); dkey=call.data.split(':',2)[2]; await state.update_data(dkey=dkey); await state.set_state(ComplaintFSM.teacher)
    await edit_or_send(call, "O‘qituvchini tanlang:", teachers_kb(dkey,'complaint:teacher','complaint:start'))
@dp.callback_query(F.data.startswith('complaint:teacher:'))
async def complaint_teacher(call:CallbackQuery,state:FSMContext):
    await answer_cb(call); _,_,dkey,tkey=call.data.split(':',3); await state.update_data(dkey=dkey,tkey=tkey); await state.set_state(ComplaintFSM.text)
    await ask_input(call, f"<b>{safe(teacher_name(dkey,tkey))}</b> ustidan shikoyat matnini yozing:")
@dp.message(ComplaintFSM.text)
async def complaint_text(m:Message,state:FSMContext):
    text=(m.text or '').strip()
    if not text:
        return await m.answer("Iltimos, shikoyat matnini oddiy xabar ko‘rinishida yozing.", reply_markup=cancel_kb())
    if len(text)<5:
        return await m.answer("Shikoyat matni juda qisqa. Iltimos, batafsilroq yozing.", reply_markup=cancel_kb())
    data=await state.get_data()
    dkey=data.get('dkey'); tkey=data.get('tkey')
    if not dkey or not tkey:
        await state.clear()
        await m.answer("Jarayonda xatolik bo‘ldi. Iltimos, qaytadan urinib ko‘ring.", reply_markup=ReplyKeyboardRemove())
        return await m.answer(home_text(), reply_markup=main_kb())
    u=get_user(m.from_user.id)
    fullname = u['fullname'] if u and 'fullname' in u.keys() else (m.from_user.full_name or '')
    phone = u['phone'] if u and 'phone' in u.keys() else ''
    username = u['username'] if u and 'username' in u.keys() else (m.from_user.username or '')
    try:
        conn.execute("""INSERT INTO complaints(user_id,teacher_id,department_key,teacher_key,fullname,phone,username,complaint_text,message_text,created_at,status)
            VALUES(?,?,?,?,?,?,?,?,?,?,?)""", (m.from_user.id,tkey,dkey,tkey,fullname,phone,username,text,text,now_str(),'Yangi'))
        conn.commit()
    except Exception:
        log.exception("Complaint save failed")
        await state.clear()
        await m.answer("Shikoyatni saqlashda xatolik bo‘ldi. Iltimos, keyinroq qayta urinib ko‘ring.", reply_markup=ReplyKeyboardRemove())
        return await m.answer(home_text(), reply_markup=main_kb())
    await state.clear()
    await m.answer("✅ Shikoyatingiz saqlandi. Tez orada ko‘rib chiqiladi.", reply_markup=ReplyKeyboardRemove())
    await m.answer(home_text(), reply_markup=main_kb())
    admin_note = (
        f"📢 Yangi shikoyat\n"
        f"O‘qituvchi: {safe(teacher_name(dkey, tkey))}\n"
        f"Kafedra: {safe(dep_name(dkey))}\n"
        f"Foydalanuvchi: {safe(fullname)} | {safe(phone)}\n\n"
        f"{safe(text[:1000])}"
    )
    asyncio.create_task(notify_admins(admin_note))

@dp.callback_query(F.data == 'suggestion:start')
async def suggestion_start(call:CallbackQuery,state:FSMContext):
    await answer_cb(call)
    await state.clear()
    if not user_registered(call.from_user.id):
        return await start_registration(call,state,'suggestion')
    await state.set_state(SuggestionFSM.text)
    await ask_input(call, "💡 Taklifingizni yozing. Matn aniq va lo‘nda bo‘lsa, ko‘rib chiqish oson bo‘ladi:")

@dp.message(SuggestionFSM.text)
async def suggestion_text(m:Message,state:FSMContext):
    text=(m.text or '').strip()
    if len(text)<5:
        return await m.answer("Taklif matni juda qisqa. Iltimos, biroz batafsilroq yozing.", reply_markup=cancel_kb())
    u=get_user(m.from_user.id)
    conn.execute("""INSERT INTO suggestions(user_id,fullname,phone,username,suggestion_text,created_at,status)
        VALUES(?,?,?,?,?,?,?)""", (m.from_user.id,u['fullname'] if u else '',u['phone'] if u else '',u['username'] if u else '',text,now_str(),'Yangi'))
    conn.commit()
    await state.clear()
    await m.answer("✅ Taklifingiz qabul qilindi. Rahmat!", reply_markup=ReplyKeyboardRemove())
    await m.answer(home_text(), reply_markup=main_kb())
    admin_note = (
        f"💡 Yangi taklif\n"
        f"Foydalanuvchi: {safe(u['fullname'] if u else '')} | {safe(u['phone'] if u else '')}\n\n"
        f"{safe(text[:1000])}"
    )
    asyncio.create_task(notify_admins(admin_note))

# =====================================================
# ADMIN HANDLERS
# =====================================================
def admin_required(func):
    @wraps(func)
    async def wrapper(event, *args, **kwargs):
        uid=event.from_user.id
        if not is_admin(uid):
            if isinstance(event, CallbackQuery): await event.answer('Ruxsat yo‘q.', show_alert=True)
            else: await event.answer('Ruxsat yo‘q.')
            return
        return await func(event,*args,**kwargs)
    return wrapper

@dp.callback_query(F.data == 'admin:menu')
@admin_required
async def admin_menu(call:CallbackQuery): await answer_cb(call); await edit_or_send(call,"<b>Admin panel</b>",admin_kb())
@dp.callback_query(F.data == 'admin:dash')
@admin_required
async def admin_dash(call:CallbackQuery):
    await answer_cb(call)
    users=conn.execute('SELECT COUNT(*) FROM users').fetchone()[0]; ratings=conn.execute('SELECT COUNT(*) FROM ratings').fetchone()[0]; comps=conn.execute('SELECT COUNT(*) FROM complaints').fetchone()[0]; suggs=conn.execute('SELECT COUNT(*) FROM suggestions').fetchone()[0]
    txt=f"📊 <b>Dashboard</b>\n\nFoydalanuvchilar: <b>{users}</b>\nBaholar: <b>{ratings}</b>\nShikoyatlar: <b>{comps}</b>\nTakliflar: <b>{suggs}</b>\nBaholash: <b>{'Ochiq' if setting_bool('voting_open') else 'Yopiq'}</b>"
    await edit_or_send(call,txt,ik([[('⬅️ Admin','admin:menu')]]))
@dp.callback_query(F.data == 'admin:teachers')
@admin_required
async def adm_teachers(call:CallbackQuery): await answer_cb(call); await edit_or_send(call,"👨‍🏫 <b>O‘qituvchilar</b>",admin_teachers_kb())
@dp.callback_query(F.data == 'tm:list')
@admin_required
async def tm_list(call:CallbackQuery): await answer_cb(call); await edit_or_send(call,"Kafedrani tanlang:",deps_kb('tm:listdep','admin:teachers'))
@dp.callback_query(F.data.startswith('tm:listdep:'))
@admin_required
async def tm_listdep(call:CallbackQuery):
    await answer_cb(call); dkey=call.data.split(':',2)[2]
    rows=[]
    for t in teachers(dkey,False):
        icon='✅' if t['is_active'] else '❌'
        rows.append([(f"{icon} {t['name'][:35]}", f"tm:detail:{dkey}:{t['teacher_key']}")])
    rows.append([('⬅️ Orqaga','tm:list')])
    await edit_or_send(call,f"<b>{safe(dep_name(dkey))}</b>\nO‘qituvchini tanlang:",InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text=a, callback_data=b) for a,b in r] for r in rows]))

@dp.callback_query(F.data.startswith('tm:detail:'))
@admin_required
async def tm_detail(call:CallbackQuery):
    await answer_cb(call); _,_,dkey,tkey=call.data.split(':',3)
    st=teacher_stats(dkey,tkey); active=conn.execute("SELECT is_active FROM teachers WHERE department_key=? AND teacher_key=?",(dkey,tkey)).fetchone()[0]
    txt=f"<b>{safe(teacher_name(dkey,tkey))}</b>\nKafedra: {safe(dep_name(dkey))}\nHolat: {'Faol' if active else 'Arxiv'}\nO‘quvchilar soni: <b>{st['student_count']}</b>\nOvozlar: <b>{st['total']}</b>\nO‘rtacha: <b>{st['avg']:.2f}</b>"
    kb=ik([[('✏️ Ism','tm:editname:'+dkey+':'+tkey),('👥 O‘quvchi','tm:editstudents:'+dkey+':'+tkey)],[('📊 Statistika','stats:teacher:'+dkey+':'+tkey),('🗑 Arxivlash','del_teacher:'+dkey+':'+tkey)],[('♻️ Tiklash','restore_teacher:'+dkey+':'+tkey)],[('⬅️ Ro‘yxat','tm:listdep:'+dkey)]])
    await edit_or_send(call,txt,kb)
@dp.callback_query(F.data == 'tm:add')
@admin_required
async def tm_add(call:CallbackQuery,state:FSMContext): await answer_cb(call); await state.set_state(TeacherFSM.add_department); await edit_or_send(call,"Qaysi kafedraga qo‘shiladi?",deps_kb('tm:adddep','admin:teachers'))
@dp.callback_query(F.data.startswith('tm:adddep:'))
@admin_required
async def tm_adddep(call:CallbackQuery,state:FSMContext): await answer_cb(call); dkey=call.data.split(':',2)[2]; await state.update_data(dkey=dkey); await state.set_state(TeacherFSM.add_name); await ask_input(call,"Yangi o‘qituvchi F.I.Sh ni kiriting:")
@dp.message(TeacherFSM.add_name)
@admin_required
async def tm_addname(m:Message,state:FSMContext):
    data=await state.get_data(); name=(m.text or '').strip()
    if len(name)<3: return await m.answer('F.I.Sh juda qisqa.', reply_markup=cancel_kb())
    key=slug(name,'t')
    conn.execute("INSERT OR IGNORE INTO teachers(teacher_key,department_key,name,created_at) VALUES(?,?,?,?)",(key,data['dkey'],name,now_str())); conn.commit(); await state.clear(); await m.answer("✅ O‘qituvchi qo‘shildi.",reply_markup=admin_teachers_kb())
@dp.callback_query(F.data == 'tm:students')
@admin_required
async def tm_students(call:CallbackQuery): await answer_cb(call); await edit_or_send(call,"Kafedrani tanlang:",deps_kb('tm:studentsdep','admin:teachers'))
@dp.callback_query(F.data.startswith('tm:studentsdep:'))
@admin_required
async def tm_studentsdep(call:CallbackQuery): await answer_cb(call); dkey=call.data.split(':',2)[2]; await edit_or_send(call,"O‘qituvchini tanlang:",teachers_kb(dkey,'tm:studentteacher','tm:students'))
@dp.callback_query(F.data.startswith('tm:studentteacher:'))
@admin_required
async def tm_studentteacher(call:CallbackQuery,state:FSMContext):
    await answer_cb(call); _,_,dkey,tkey=call.data.split(':',3); await state.update_data(dkey=dkey,tkey=tkey); await state.set_state(TeacherFSM.student_count)
    st=teacher_stats(dkey,tkey); action='✏️ O‘zgartirish' if st['student_count'] else '➕ Kiritish'
    await ask_input(call,f"{safe(teacher_name(dkey,tkey))}\nHozirgi o‘quvchilar soni: <b>{st['student_count']}</b>\n\n{action}: yangi sonni kiriting.")
@dp.message(TeacherFSM.student_count)
@admin_required
async def tm_student_save(m:Message,state:FSMContext):
    if not (m.text or "").isdigit(): return await m.answer("Faqat raqam kiriting.", reply_markup=cancel_kb())
    data=await state.get_data(); conn.execute("UPDATE teachers SET student_count=? WHERE department_key=? AND teacher_key=?",(int(m.text),data['dkey'],data['tkey'])); conn.commit(); await state.clear(); await m.answer("✅ Saqlandi.",reply_markup=admin_teachers_kb())
@dp.callback_query(F.data == 'tm:search')
@admin_required
async def tm_search(call:CallbackQuery,state:FSMContext): await answer_cb(call); await state.set_state(TeacherFSM.search); await ask_input(call,"Qidirish uchun ism/familiya kiriting:")
@dp.message(TeacherFSM.search)
@admin_required
async def tm_search_msg(m:Message,state:FSMContext):
    query=(m.text or "").strip()
    if len(query) < 2: return await m.answer("Qidiruv so‘zi juda qisqa.", reply_markup=cancel_kb())
    q=f"%{query}%"; rows=conn.execute("SELECT t.*,d.name dep FROM teachers t JOIN departments d ON d.department_key=t.department_key WHERE t.name LIKE ? LIMIT 30",(q,)).fetchall(); await state.clear()
    if not rows: return await m.answer("Topilmadi.",reply_markup=admin_teachers_kb())
    await m.answer('\n\n'.join([f"<b>{safe(r['name'])}</b>\n{safe(r['dep'])}\nO‘quvchi: {r['student_count'] or 0}" for r in rows]),reply_markup=admin_teachers_kb())

@dp.callback_query(F.data == 'admin:tstats')
@admin_required
async def adm_tstats(call:CallbackQuery): await answer_cb(call); await edit_or_send(call,"Kafedrani tanlang:",deps_kb('stats:dep','admin:menu'))
@dp.callback_query(F.data.startswith('stats:dep:'))
@admin_required
async def stats_dep(call:CallbackQuery): await answer_cb(call); dkey=call.data.split(':',2)[2]; await edit_or_send(call,"O‘qituvchini tanlang:",teachers_kb(dkey,'stats:teacher','admin:tstats'))
@dp.callback_query(F.data.startswith('stats:teacher:'))
@admin_required
async def stats_teacher(call:CallbackQuery):
    await answer_cb(call); _,_,dkey,tkey=call.data.split(':',3)
    kb=ik([[('📢 Shikoyatlar',f'complaints:teacher:{dkey}:{tkey}')],[('✏️ Tahrirlash',f'edit_teacher:{dkey}:{tkey}'),('🗑 O‘chirish',f'del_teacher:{dkey}:{tkey}')],[('⬅️ Orqaga',f'stats:dep:{dkey}')]])
    await edit_or_send(call,teacher_stats_text(dkey,tkey),kb)
@dp.callback_query(F.data.startswith('complaints:teacher:'))
@admin_required
async def complaints_teacher(call:CallbackQuery):
    await answer_cb(call); _,_,dkey,tkey=call.data.split(':',3)
    rows=conn.execute("SELECT * FROM complaints WHERE department_key=? AND teacher_key=? ORDER BY id DESC LIMIT 20",(dkey,tkey)).fetchall()
    txt=f"📢 <b>{safe(teacher_name(dkey,tkey))} ustidan shikoyatlar</b>\n\n" + ('Hali mavjud emas.' if not rows else '\n\n'.join([f"#{r['id']} • {safe(r['status'])} • {safe(r['created_at'])}\n{safe(r['complaint_text'] or r['message_text'])}" for r in rows]))
    await edit_or_send(call,txt[:4000],ik([[('⬅️ Orqaga',f'stats:teacher:{dkey}:{tkey}')]]))
@dp.callback_query(F.data.startswith('del_teacher:'))
@admin_required
async def del_teacher(call:CallbackQuery): await answer_cb(call); _,dkey,tkey=call.data.split(':',2); conn.execute("UPDATE teachers SET is_active=0 WHERE department_key=? AND teacher_key=?",(dkey,tkey)); conn.commit(); await edit_or_send(call,"✅ O‘qituvchi o‘chirildi (arxivlandi).",admin_teachers_kb())


@dp.callback_query(F.data.startswith('tm:editname:'))
@admin_required
async def tm_edit_name_start(call:CallbackQuery,state:FSMContext):
    await answer_cb(call); _,_,dkey,tkey=call.data.split(':',3)
    await state.update_data(dkey=dkey,tkey=tkey); await state.set_state(TeacherEditFSM.name)
    await ask_input(call,f"Hozirgi nom: <b>{safe(teacher_name(dkey,tkey))}</b>\n\nYangi F.I.Sh ni kiriting:")

@dp.message(TeacherEditFSM.name)
@admin_required
async def tm_edit_name_save(m:Message,state:FSMContext):
    name=(m.text or '').strip()
    if len(name)<3: return await m.answer("Nom juda qisqa.", reply_markup=cancel_kb())
    data=await state.get_data(); conn.execute("UPDATE teachers SET name=? WHERE department_key=? AND teacher_key=?",(name,data['dkey'],data['tkey'])); conn.commit(); await state.clear()
    await m.answer("✅ O‘qituvchi nomi yangilandi.", reply_markup=admin_teachers_kb())

@dp.callback_query(F.data.startswith('tm:editstudents:'))
@admin_required
async def tm_edit_students_start(call:CallbackQuery,state:FSMContext):
    await answer_cb(call); _,_,dkey,tkey=call.data.split(':',3)
    await state.update_data(dkey=dkey,tkey=tkey); await state.set_state(TeacherEditFSM.student_count)
    await ask_input(call,f"{safe(teacher_name(dkey,tkey))}\nYangi o‘quvchilar sonini kiriting:")

@dp.message(TeacherEditFSM.student_count)
@admin_required
async def tm_edit_students_save(m:Message,state:FSMContext):
    if not (m.text or '').isdigit(): return await m.answer("Faqat raqam kiriting.", reply_markup=cancel_kb())
    data=await state.get_data(); conn.execute("UPDATE teachers SET student_count=? WHERE department_key=? AND teacher_key=?",(int(m.text),data['dkey'],data['tkey'])); conn.commit(); await state.clear()
    await m.answer("✅ O‘quvchilar soni yangilandi.", reply_markup=admin_teachers_kb())

@dp.callback_query(F.data.startswith('edit_teacher:'))
@admin_required
async def edit_teacher_alias(call:CallbackQuery,state:FSMContext):
    await answer_cb(call); _,dkey,tkey=call.data.split(':',2)
    await state.update_data(dkey=dkey,tkey=tkey); await state.set_state(TeacherEditFSM.name)
    await ask_input(call,f"Hozirgi nom: <b>{safe(teacher_name(dkey,tkey))}</b>\n\nYangi F.I.Sh ni kiriting:")

@dp.callback_query(F.data.startswith('restore_teacher:'))
@admin_required
async def restore_teacher(call:CallbackQuery):
    await answer_cb(call); _,dkey,tkey=call.data.split(':',2)
    conn.execute("UPDATE teachers SET is_active=1 WHERE department_key=? AND teacher_key=?",(dkey,tkey)); conn.commit()
    await edit_or_send(call,"✅ O‘qituvchi tiklandi.",admin_teachers_kb())

@dp.callback_query(F.data == 'dep:manage')
@admin_required
async def dep_manage(call:CallbackQuery):
    await answer_cb(call)
    await edit_or_send(call,"🏛 <b>Kafedralar boshqaruvi</b>",ik([[('➕ Qo‘shish','dep:add'),('📋 Ro‘yxat','dep:list')],[('⬅️ Admin','admin:menu')]]))

@dp.callback_query(F.data == 'dep:list')
@admin_required
async def dep_list(call:CallbackQuery):
    await answer_cb(call)
    rows=[[(('✅ ' if d['is_active'] else '❌ ')+d['name'][:35], f"dep:detail:{d['department_key']}")] for d in departments(False)]
    rows.append([('⬅️ Orqaga','dep:manage')])
    await edit_or_send(call,"Kafedrani tanlang:",InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text=a, callback_data=b) for a,b in r] for r in rows]))

@dp.callback_query(F.data.startswith('dep:detail:'))
@admin_required
async def dep_detail(call:CallbackQuery):
    await answer_cb(call); dkey=call.data.split(':',2)[2]
    d=conn.execute("SELECT * FROM departments WHERE department_key=?",(dkey,)).fetchone(); st=department_stats(dkey)
    txt=f"<b>{safe(d['name'])}</b>\nHolat: {'Faol' if d['is_active'] else 'Arxiv'}\nO‘qituvchilar: <b>{st['teachers']}</b>\nBaholar: <b>{st['votes']}</b>\nFinal score: <b>{st['final']:.3f}</b>"
    await edit_or_send(call,txt,ik([[('✏️ Nom','dep:edit:'+dkey),('🗑 Arxiv','dep:archive:'+dkey)],[('♻️ Tiklash','dep:restore:'+dkey)],[('⬅️ Ro‘yxat','dep:list')]]))

@dp.callback_query(F.data == 'dep:add')
@admin_required
async def dep_add_start(call:CallbackQuery,state:FSMContext):
    await answer_cb(call); await state.set_state(DepartmentFSM.add_name); await ask_input(call,"Yangi kafedra nomini kiriting:")

@dp.message(DepartmentFSM.add_name)
@admin_required
async def dep_add_save(m:Message,state:FSMContext):
    name=(m.text or '').strip()
    if len(name)<3: return await m.answer("Nom juda qisqa.", reply_markup=cancel_kb())
    key=slug(name,'d'); order=conn.execute("SELECT COALESCE(MAX(sort_order),0)+1 FROM departments").fetchone()[0]
    conn.execute("INSERT INTO departments(department_key,name,sort_order,created_at) VALUES(?,?,?,?)",(key,name,order,now_str())); conn.commit(); await state.clear()
    await m.answer("✅ Kafedra qo‘shildi.", reply_markup=admin_teachers_kb())

@dp.callback_query(F.data.startswith('dep:edit:'))
@admin_required
async def dep_edit_start(call:CallbackQuery,state:FSMContext):
    await answer_cb(call); dkey=call.data.split(':',2)[2]; await state.update_data(dkey=dkey); await state.set_state(DepartmentFSM.edit_name)
    await ask_input(call,f"Hozirgi nom: <b>{safe(dep_name(dkey))}</b>\nYangi nomni kiriting:")

@dp.message(DepartmentFSM.edit_name)
@admin_required
async def dep_edit_save(m:Message,state:FSMContext):
    name=(m.text or '').strip()
    if len(name)<3: return await m.answer("Nom juda qisqa.", reply_markup=cancel_kb())
    data=await state.get_data(); conn.execute("UPDATE departments SET name=? WHERE department_key=?",(name,data['dkey'])); conn.commit(); await state.clear()
    await m.answer("✅ Kafedra yangilandi.", reply_markup=admin_teachers_kb())

@dp.callback_query(F.data.startswith('dep:archive:'))
@admin_required
async def dep_archive(call:CallbackQuery):
    await answer_cb(call); dkey=call.data.split(':',2)[2]; conn.execute("UPDATE departments SET is_active=0 WHERE department_key=?",(dkey,)); conn.commit(); await edit_or_send(call,"✅ Kafedra arxivlandi.",admin_teachers_kb())

@dp.callback_query(F.data.startswith('dep:restore:'))
@admin_required
async def dep_restore(call:CallbackQuery):
    await answer_cb(call); dkey=call.data.split(':',2)[2]; conn.execute("UPDATE departments SET is_active=1 WHERE department_key=?",(dkey,)); conn.commit(); await edit_or_send(call,"✅ Kafedra tiklandi.",admin_teachers_kb())

@dp.callback_query(F.data.startswith('complaint:status:'))
@admin_required
async def complaint_status(call:CallbackQuery):
    await answer_cb(call); _,_,cid,status=call.data.split(':',3)
    status_map={'new':'Yangi','process':'Ko‘rib chiqilmoqda','closed':'Yopilgan'}
    conn.execute("UPDATE complaints SET status=? WHERE id=?",(status_map.get(status,'Yangi'),cid)); conn.commit()
    await edit_or_send(call,"✅ Shikoyat statusi yangilandi.",ik([[('📢 Shikoyatlar','admin:complaints')],[('⬅️ Admin','admin:menu')]]))

@dp.callback_query(F.data == 'admin:dept_rating')
@admin_required
async def adm_dep_rating(call:CallbackQuery): await answer_cb(call); await edit_or_send(call,departments_rating_text(),ik([[('⬅️ Admin','admin:menu')]]))
@dp.callback_query(F.data == 'admin:complaints')
@admin_required
async def adm_complaints(call:CallbackQuery):
    await answer_cb(call); rows=conn.execute("SELECT * FROM complaints ORDER BY id DESC LIMIT 30").fetchall()
    if not rows:
        return await edit_or_send(call,"📢 <b>Shikoyatlar</b>\n\nMavjud emas.",ik([[('⬅️ Admin','admin:menu')]]))
    kb_rows=[[(f"#{r['id']} • {r['status']} • {teacher_name(r['department_key'],r['teacher_key'])[:24]}", f"complaint:view:{r['id']}")] for r in rows]
    kb_rows.append([('⬅️ Admin','admin:menu')])
    await edit_or_send(call,"📢 <b>Shikoyatlar</b>\nBittasini tanlang:",InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text=a, callback_data=b) for a,b in row] for row in kb_rows]))

@dp.callback_query(F.data.startswith('complaint:view:'))
@admin_required
async def complaint_view(call:CallbackQuery):
    await answer_cb(call); cid=call.data.split(':',2)[2]
    r=conn.execute("SELECT * FROM complaints WHERE id=?",(cid,)).fetchone()
    if not r: return await edit_or_send(call,"Topilmadi.",ik([[('⬅️ Orqaga','admin:complaints')]]))
    txt=f"📢 <b>Shikoyat #{r['id']}</b>\nStatus: <b>{safe(r['status'])}</b>\nFoydalanuvchi: {safe(r['fullname'])} | {safe(r['phone'])}\nKafedra: {safe(dep_name(r['department_key']))}\nO‘qituvchi: {safe(teacher_name(r['department_key'],r['teacher_key']))}\nSana: {safe(r['created_at'])}\n\n{safe(r['complaint_text'] or r['message_text'])}"
    kb=ik([[('Yangi','complaint:status:'+cid+':new'),('Jarayonda','complaint:status:'+cid+':process'),('Yopilgan','complaint:status:'+cid+':closed')],[('⬅️ Ro‘yxat','admin:complaints')]])
    await edit_or_send(call,txt[:4000],kb)
@dp.callback_query(F.data == 'admin:suggestions')
@admin_required
async def adm_suggestions(call:CallbackQuery):
    await answer_cb(call)
    rows=conn.execute("SELECT * FROM suggestions ORDER BY id DESC LIMIT 30").fetchall()
    if not rows:
        return await edit_or_send(call,"💡 <b>Takliflar</b>\n\nMavjud emas.",ik([[('⬅️ Admin','admin:menu')]]))
    kb_rows=[[(f"#{r['id']} • {r['status']} • {(r['fullname'] or 'Foydalanuvchi')[:24]}", f"suggestion:view:{r['id']}")] for r in rows]
    kb_rows.append([('⬅️ Admin','admin:menu')])
    await edit_or_send(call,"💡 <b>Takliflar</b>\nBittasini tanlang:",InlineKeyboardMarkup(inline_keyboard=[[InlineKeyboardButton(text=a, callback_data=b) for a,b in row] for row in kb_rows]))

@dp.callback_query(F.data.startswith('suggestion:view:'))
@admin_required
async def suggestion_view(call:CallbackQuery):
    await answer_cb(call); sid=call.data.split(':',2)[2]
    r=conn.execute("SELECT * FROM suggestions WHERE id=?",(sid,)).fetchone()
    if not r: return await edit_or_send(call,"Topilmadi.",ik([[('⬅️ Orqaga','admin:suggestions')]]))
    txt=f"💡 <b>Taklif #{r['id']}</b>\nStatus: <b>{safe(r['status'])}</b>\nFoydalanuvchi: {safe(r['fullname'])} | {safe(r['phone'])}\nSana: {safe(r['created_at'])}\n\n{safe(r['suggestion_text'])}"
    kb=ik([[('Yangi','suggestion:status:'+sid+':new'),('Jarayonda','suggestion:status:'+sid+':process'),('Yopilgan','suggestion:status:'+sid+':closed')],[('⬅️ Ro‘yxat','admin:suggestions')]])
    await edit_or_send(call,txt[:4000],kb)

@dp.callback_query(F.data.startswith('suggestion:status:'))
@admin_required
async def suggestion_status(call:CallbackQuery):
    await answer_cb(call); _,_,sid,status=call.data.split(':',3)
    status_map={'new':'Yangi','process':'Ko‘rib chiqilmoqda','closed':'Yopilgan'}
    conn.execute("UPDATE suggestions SET status=? WHERE id=?",(status_map.get(status,'Yangi'),sid)); conn.commit()
    await edit_or_send(call,"✅ Taklif statusi yangilandi.",ik([[('💡 Takliflar','admin:suggestions')],[('⬅️ Admin','admin:menu')]]))

@dp.callback_query(F.data == 'admin:rating')
@admin_required
async def admin_rating_menu(call:CallbackQuery):
    await answer_cb(call)
    await edit_or_send(call, "🏆 <b>Reyting</b>\nKerakli reyting turini tanlang:", ik([[('👨‍🏫 O‘qituvchilar','admin:rating:teachers'),('🏛 Kafedralar','admin:rating:departments')],[('⬅️ Admin','admin:menu')]]))

@dp.callback_query(F.data == 'admin:rating:teachers')
@admin_required
async def admin_rating_teachers(call:CallbackQuery):
    await answer_cb(call)
    await edit_or_send(call, admin_teacher_top10_text(), ik([[('📥 Excel export','export:rating_teachers')],[('⬅️ Reyting','admin:rating')]]))

@dp.callback_query(F.data == 'admin:rating:departments')
@admin_required
async def admin_rating_departments(call:CallbackQuery):
    await answer_cb(call)
    await edit_or_send(call, admin_department_rating_text(), ik([[('📥 Excel export','export:rating_departments')],[('⬅️ Reyting','admin:rating')]]))

@dp.callback_query(F.data == 'export:rating_teachers')
@admin_required
async def export_rating_teachers_cb(call:CallbackQuery):
    await answer_cb(call,'Tayyorlanmoqda...')
    path=export_teacher_ranking_excel()
    await call.message.answer_document(FSInputFile(path), caption='TOP 10 o‘qituvchilar reytingi Excel tayyor.')

@dp.callback_query(F.data == 'export:rating_departments')
@admin_required
async def export_rating_departments_cb(call:CallbackQuery):
    await answer_cb(call,'Tayyorlanmoqda...')
    path=export_department_ranking_excel()
    await call.message.answer_document(FSInputFile(path), caption='Kafedralar reytingi Excel tayyor.')

@dp.callback_query(F.data == 'admin:export')
@admin_required
async def adm_export(call:CallbackQuery): await answer_cb(call); await edit_or_send(call,"📁 <b>Export / Backup</b>",ik([[('Excel export','export:excel'),('Backup ZIP','export:backup')],[('⬅️ Admin','admin:menu')]]))
@dp.callback_query(F.data == 'export:excel')
@admin_required
async def export_excel_cb(call:CallbackQuery):
    await answer_cb(call,'Tayyorlanmoqda...'); path=export_all_excel(); await call.message.answer_document(FSInputFile(path), caption='Excel export tayyor.')
@dp.callback_query(F.data == 'export:backup')
@admin_required
async def export_backup_cb(call:CallbackQuery):
    await answer_cb(call,'Backup tayyorlanmoqda...'); path=make_backup(); await call.message.answer_document(FSInputFile(path), caption='Backup ZIP tayyor.')
@dp.callback_query(F.data == 'admin:settings')
@admin_required
async def adm_settings(call:CallbackQuery): await answer_cb(call); await edit_or_send(call,"⚙️ <b>Sozlamalar</b>",admin_settings_kb())
@dp.callback_query(F.data == 'set:voting')
@admin_required
async def set_voting(call:CallbackQuery):
    await answer_cb(call); cur=setting_bool('voting_open'); set_setting('voting_open','0' if cur else '1')
    await edit_or_send(call,f"Baholash holati: <b>{'Ochiq' if not cur else 'Yopiq'}</b>",admin_settings_kb())
@dp.callback_query(F.data == 'set:deps')
@admin_required
async def set_deps(call:CallbackQuery): await dep_manage(call)

@dp.callback_query(F.data == 'set:complaints')
@admin_required
async def set_complaints(call:CallbackQuery): await adm_complaints(call)

@dp.callback_query(F.data == 'set:rating')
@admin_required
async def set_rating(call:CallbackQuery):
    await admin_rating_menu(call)

@dp.callback_query(F.data == 'set:backup')
@admin_required
async def set_backup(call:CallbackQuery):
    await answer_cb(call); cur=setting_bool('daily_backup');
    await edit_or_send(call,f"📦 Backup sozlamalari\nAvtomatik backup: <b>{'Yoqilgan' if cur else 'O‘chirilgan'}</b>",ik([[('On/Off','backup:toggle'),('Manual backup','export:backup')],[('⬅️ Sozlamalar','admin:settings')]]))

@dp.callback_query(F.data == 'backup:toggle')
@admin_required
async def backup_toggle(call:CallbackQuery):
    await answer_cb(call); cur=setting_bool('daily_backup'); set_setting('daily_backup','0' if cur else '1'); await set_backup(call)

@dp.callback_query(F.data == 'set:admins')
@admin_required
async def set_admins(call:CallbackQuery):
    await answer_cb(call); extra=get_setting('extra_admins','') or 'yo‘q'
    await edit_or_send(call,f"🔐 Adminlar\nAsosiy ADMIN_IDS: <code>{','.join(map(str,ADMIN_IDS))}</code>\nQo‘shimcha: <code>{safe(extra)}</code>",ik([[('➕ Admin qo‘shish','admin:add'),('➖ Admin olish','admin:remove')],[('⬅️ Sozlamalar','admin:settings')]]))

@dp.callback_query(F.data == 'admin:add')
@admin_required
async def admin_add_start(call:CallbackQuery,state:FSMContext):
    await answer_cb(call); await state.set_state(AdminFSM.admin_add); await ask_input(call,"Yangi admin Telegram ID raqamini kiriting:")

@dp.message(AdminFSM.admin_add)
@admin_required
async def admin_add_save(m:Message,state:FSMContext):
    if not (m.text or '').strip().isdigit(): return await m.answer("Faqat Telegram ID raqam kiriting.", reply_markup=cancel_kb())
    ids=set(filter(None,get_setting('extra_admins','').split(','))); ids.add(m.text.strip()); set_setting('extra_admins',','.join(sorted(ids))); await state.clear(); await m.answer("✅ Admin qo‘shildi.", reply_markup=admin_kb())

@dp.callback_query(F.data == 'admin:remove')
@admin_required
async def admin_remove_start(call:CallbackQuery,state:FSMContext):
    await answer_cb(call); await state.set_state(AdminFSM.admin_remove); await ask_input(call,"Olib tashlanadigan qo‘shimcha admin ID raqamini kiriting:")

@dp.message(AdminFSM.admin_remove)
@admin_required
async def admin_remove_save(m:Message,state:FSMContext):
    ids=set(filter(None,get_setting('extra_admins','').split(','))); ids.discard((m.text or '').strip()); set_setting('extra_admins',','.join(sorted(ids))); await state.clear(); await m.answer("✅ Admin ro‘yxati yangilandi.", reply_markup=admin_kb())

# legacy/admin export commands
@dp.message(Command('backup'))
@admin_required
async def backup_cmd(m:Message): path=make_backup(); await m.answer_document(FSInputFile(path), caption='Backup ZIP')
@dp.message(Command('export'))
@admin_required
async def export_cmd(m:Message): path=export_all_excel(); await m.answer_document(FSInputFile(path), caption='Excel export')

# fallback buttons from old bot
@dp.message(F.text.in_({'📝 O‘qituvchini baholash','📝 Baholash','🗳 Ovoz berish','Ovoz berish','Baholash'}))
async def msg_rate(m:Message,state:FSMContext):
    await state.clear(); ensure_user_obj(m)
    if not setting_bool('voting_open','1'):
        return await m.answer('Baholash hozircha yopilgan.')
    await start_registration(m,state,'rate') if not user_registered(m.from_user.id) else await m.answer(MOTIVATION,reply_markup=deps_kb('rate:dep','home'))
@dp.message(F.text.in_({'📢 Shikoyat yuborish','📢 Shikoyat','Shikoyat yuborish'}))
async def msg_complaint(m:Message,state:FSMContext):
    await state.clear(); ensure_user_obj(m)
    await start_registration(m,state,'complaint') if not user_registered(m.from_user.id) else await m.answer('Kafedrani tanlang:',reply_markup=deps_kb('complaint:dep','home'))

@dp.message(F.text.in_({'💡 Taklif yuborish','💡 Taklif','Taklif yuborish','Takliflar'}))
async def msg_suggestion(m:Message,state:FSMContext):
    await state.clear(); ensure_user_obj(m)
    if not user_registered(m.from_user.id):
        return await start_registration(m,state,'suggestion')
    await state.set_state(SuggestionFSM.text)
    await ask_input(m, '💡 Taklifingizni yozing:')


@dp.message(Command('cancel'))
async def cancel_cmd(m: Message, state: FSMContext):
    await state.clear()
    await m.answer('Jarayon bekor qilindi.', reply_markup=ReplyKeyboardRemove())
    await m.answer(home_text(), reply_markup=main_kb())

@dp.callback_query(F.data == 'admin:settings')
@admin_required
async def adm_settings_duplicate_guard(call:CallbackQuery):
    # This guard is intentionally unreachable if the main settings handler is registered;
    # kept only for old deployments with partial reloads.
    await answer_cb(call)
    await edit_or_send(call,"⚙️ <b>Sozlamalar</b>",admin_settings_kb())

@dp.callback_query(F.data.startswith('set:'))
@admin_required
async def unknown_settings_callback(call: CallbackQuery):
    await answer_cb(call)
    await edit_or_send(call, "⚙️ <b>Sozlamalar</b>\nBu tugma uchun sozlamalar oynasi qayta ochildi.", admin_settings_kb())

@dp.callback_query(F.data.startswith('admin:'))
@admin_required
async def unknown_admin_callback(call: CallbackQuery):
    await answer_cb(call)
    await edit_or_send(call, '<b>Admin panel</b>\nBu tugma uchun bo‘lim qayta ochildi.', admin_kb())

@dp.callback_query(F.data.startswith('rate:'))
async def unknown_rate_callback(call: CallbackQuery):
    await answer_cb(call)
    await edit_or_send(call, MOTIVATION, deps_kb('rate:dep','home'))

@dp.errors()
async def errors_handler(event):
    log.exception("Update error: %s", event.exception)
    return True

async def daily_backup_job():
    if setting_bool('daily_backup','1'):
        try: make_backup(); log.info('Daily backup created')
        except Exception: log.exception('Daily backup failed')

async def main():
    init_db()
    if AsyncIOScheduler:
        scheduler=AsyncIOScheduler(timezone='Asia/Tashkent')
        scheduler.add_job(daily_backup_job, 'cron', hour=3, minute=0)
        scheduler.start()
    log.info('Bot started')
    await dp.start_polling(bot, allowed_updates=dp.resolve_used_update_types())

if __name__ == '__main__':
    asyncio.run(main())
